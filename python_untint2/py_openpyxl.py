from openpyxl import load_workbook

wb = load_workbook('taobao1.xlsx')
ws = wb.active
cell_value= ws['A1']
print(cell_value)
print(cell_value.value)
#cell_value.value = '新写入'
#print(cell_value.value)
cell_value = ws.cell(1, 1)
print(cell_value)
print(ws.max_row)
cell_max = ws.max_row + 1
ws.cell(cell_max, 1).value = '新添加'
print(ws.cell((ws.max_row-1), 1).value)
print(ws.max_row)
wb.save('taobao1.xlsx')