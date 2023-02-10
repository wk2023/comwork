from selenium import webdriver  # 引入浏览器驱动程序
from selenium.webdriver.chrome.options import Options
import time  # 引入时间模块
from selenium.webdriver.common.by import By  # 引入by 函数 定位用
from random import randint  # 引入随机数模块
from openpyxl import load_workbook  # 引入 excel操作库
# import os  # 引入系统模块
# 启动浏览器的命令  chrome.exe --remote-debugging-port=9222 --user-data-dir="D:\data"

# os.system('chrome.exe --remote-debugging-port=9222 --user-data-dir="D:\data"')  # 打开浏览器
wb = load_workbook('taobao.xlsx')  # 引入数据存储表格
ws = wb.active
wb_keywords = load_workbook('keywords.xlsx')
ws_keywords = wb_keywords.active
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
browser = webdriver.Chrome(options=chrome_options)  # 浏览器驱动启动
# 默认网页已经打开淘宝  否则执行下面语句
time.sleep(3)
browser.get('http://www.taobao.com')  # 网页转跳到淘宝网
time.sleep(3)
browser.find_element(By.XPATH, '/html/body/div[2]/div/div/div[2]/div/div[1]/form/div[3]/input').send_keys('坐垫')  # 淘宝搜索框 输入坐垫关键词
time.sleep(5)
browser.find_element(By.XPATH, '/html/body/div[2]/div/div/div[2]/div/div[1]/form/div[1]/button').click()  # 点击主页淘宝搜索按钮
time.sleep(5)
for words_number in range(1, ws_keywords.max_row):  # 开始循环要搜索的关键词
    keyword = ws_keywords.cell(words_number, 1).value  # 获取关键词
    browser.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div/div/div[3]/form/div/div[2]/div[1]/div/input').clear()  # 输入框清空
    browser.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div/div/div[3]/form/div/div[2]/div[1]/div/input').send_keys(keyword)  # 在搜索框输入关键词
    time.sleep(3)
    browser.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div/div/div[3]/form/button').click()  # 点击搜索按钮
    time.sleep(3)
    browser.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[16]/div/div[1]/div/ul/li[2]/a').click()  # 点击销量排序
    time.sleep(5)
    for page in range(1, 3):  # 开始 网页循环 一页页 翻 设定翻页书 特别注意range 函数取值 最后一位数取不到
        divtitles = browser.find_elements(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[21]/div/div/div[1]/div/*/*/a')  # 获取产品标题列表
        divshopname = browser.find_elements(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[21]/div/div/div[1]/div/div[2]/div[3]/div[1]/a')  # 获取厂家名称列表
        divdeal_cnt = browser.find_elements(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[21]/div/div/div[1]/div/div[2]/div[1]/div[2]')   # 获取收货人数列表
        divpice = browser.find_elements(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[21]/div/div/div[1]/div/div[2]/div[1]/div[1]/strong')  # 获取价格列表
        divcity = browser.find_elements(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[21]/div/div/div[1]/div/div[2]/div[3]/div[2]')  # 获取厂家列表
        divicon = browser.find_elements(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[21]/div/div/div[1]/div/div[2]/div[4]/div[1]/ul/li/a/span')  # 获取店铺标志列表 用来判读是不是天猫店铺
        print(divicon)
        for i in range(0, len(divtitles)):   # 循环本页数据 获取各种产品详细数据
            print(i+1)
            ws.cell((ws.max_row+1), 1).value = i+1   # 新建一行 在第一列写入当前开始数
            print(divtitles[i].get_attribute('href'))
            divtitles_href = divtitles[i].get_attribute('href')  # 获取产品链接地址
            ws.cell(ws.max_row, 3).value = divtitles_href  # 在excel表格 最后一行第三列写入产品链接地址
            print(divtitles[i].text)
            ws.cell(ws.max_row, 2).value = divtitles[i].text  # 在excel最后一行第二列写入产品标题
            print(divshopname[i].text)
            ws.cell(ws.max_row, 4).value = divshopname[i].text  # 在excel最后一行第四列写出店铺名称
            print(divshopname[i].get_attribute('href'))
            ws.cell(ws.max_row, 5).value = divshopname[i].get_attribute('href')  # 在最后一行 第五列写入店铺链接地址
            print(divdeal_cnt[i].text)
            ws.cell(ws.max_row, 6).value = divdeal_cnt[i].text  # 在最后一行第六列 写入收货人数
            print(divpice[i].text)
            ws.cell(ws.max_row, 7).value = divpice[i].text  # 在最后一行 第七列 写入价格
            print(divcity[i].text)
            ws.cell(ws.max_row, 8).value = divcity[i].text  # 在最后一页 第八列写入 厂家省份
            shop_icon = divicon[i].get_attribute('class')
            if shop_icon == 'icon-service-tianmao':
                ws.cell(ws.max_row, 9).value = '天猫店铺'
            else:
                ws.cell(ws.max_row, 9).value = '淘宝C店'
    wb.save('taobao.xlsx')  # 保存excel
    next_booten = browser.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[3]/div[1]/div[26]/div/div/div/ul/li[@class="item next"]/a')  # 获取下一页按钮
    next_booten.click()  # 点击下一页按钮
    time.sleep(randint(5, 10))  # 随机等待5到10秒
