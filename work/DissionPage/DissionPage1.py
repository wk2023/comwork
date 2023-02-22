from DrissionPage import ChromiumPage
from openpyxl import load_workbook
import time
from random import randint


# wb = load_workbook('DissionPage/taobao.xlsx')  # 引入数据存储表格
# ws = wb.active
# wb_keywords = load_workbook('U:\python_git\comwork\work\DissionPage\keywords.xlsx')
# ws_keywords = wb_keywords.active
page = ChromiumPage()
nub = 1
for pages in range(1,50):
    lists = page.eles('.item J_MouserOnverReq  ')
    shoplink = page.eles('.shopname J_MouseEneterLeave J_ShopInfo')
    items = page.eles('.J_ClickStat')   
    for i in range(0, len(lists)):
        lists_text = lists[i].text
        shoplink_link = shoplink[i].attr('href')
        item_id = items[i].attr('data-nid')
        item_link = items[i].attr('href')
        current_time = int(time.time())
        print(current_time) # 1631186249
        # 转换为localtime
        localtime = time.localtime(current_time)
        # 利用strftime()函数重新格式化时间
        dt = time.strftime('%Y:%m:%d/%H:%M:%S', localtime)
        all_list =str(nub)+' '+ dt + ' '+lists_text + ' '+shoplink_link + ' '+item_id + ' '+item_link
        if 'tmall' in item_link:
            all_list = all_list +' '+"天猫"
        else:
            all_list = all_list +' '+"淘宝"
        all = all_list.split()
        print(all)
        nub +=1
    page.ele('.J_Ajax num icon-tag').click()
    time.sleep(10)