from openpyxl import load_workbook

wb = load_workbook('taobao1.xlsx')
print(wb)

ws = wb.active  # 获取当前活跃的worksheet对象（sheet表）
print(ws)

cell = ws['A1']  # 获取指定位置的单元格对象
print(cell)

# 获取指定单元格的值
print(ws['A1'].value)
print(ws.cell(1, 1).value)